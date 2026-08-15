#!/usr/bin/env python3
"""Discover possible software vendors from public-sector multiplier sources.

This is the first stage before the provider crawler. It reads fairs,
conferences and platform sources, crawls their public pages, extracts possible
vendor links, and optionally appends strong external company candidates to the
curation seed inbox.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from crawl4ai import AsyncWebCrawler, BrowserConfig, CacheMode, CrawlerRunConfig
from openpyxl import load_workbook


DEFAULT_MULTIPLIERS = "data/crawler/discovery/multiplier-sources.json"
DEFAULT_CURATION = "data/crawler/discovery/curation-seeds.json"
DEFAULT_EXCEL = "Amtshelden_Zielkundenliste_Sponsoring_2026 (1).xlsx"
DEFAULT_OUT = "data/crawler/discovery/runs"

COMPANY_SUFFIX_RE = re.compile(
    r"\b(gmbh|ag|se|kg|gbr|ug|inc|llc|ltd|group|systems|software|solutions|consulting|technologies|digital)\b",
    re.IGNORECASE,
)

NOISE_HOSTS = {
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "x.com",
    "twitter.com",
    "youtube.com",
    "youtu.be",
    "google.com",
    "google.de",
    "maps.google.com",
    "apple.com",
    "microsoft.com",
    "adobe.com",
    "vimeo.com",
    "eventbrite.de",
    "eventbrite.com",
    "xing.com",
    "pretix.eu",
    "doo.net",
}

NOISE_PATH_TERMS = [
    "impressum",
    "datenschutz",
    "privacy",
    "cookie",
    "agb",
    "terms",
    "kontakt",
    "contact",
    "newsletter",
    "ticket",
    "login",
    "register",
    "anmeldung",
    "presse",
    "press",
    "hotel",
    "anfahrt",
    "arrival",
    "venue",
]

DISCOVERY_PAGE_TERMS = [
    "aussteller",
    "ausstellerliste",
    "exhibitor",
    "exhibitors",
    "partner",
    "sponsor",
    "sponsoren",
    "marktplatz",
    "marketplace",
    "teilnehmer",
    "companies",
    "unternehmen",
    "programm",
    "agenda",
]

PROFILE_PAGE_TERMS = [
    "aussteller",
    "exhibitor",
    "partner",
    "sponsor",
    "company",
    "unternehmen",
]

FIT_TERMS = {
    "public_sector": [
        "behoerde",
        "behörde",
        "verwaltung",
        "kommune",
        "kommunal",
        "public sector",
        "egovernment",
        "e-government",
        "govtech",
        "ozg",
    ],
    "software": [
        "software",
        "plattform",
        "platform",
        "saas",
        "cloud",
        "ki",
        "ai",
        "digital",
        "daten",
        "security",
        "cyber",
        "prozess",
        "workflow",
        "portal",
        "fachverfahren",
        "dokument",
    ],
}

STRONG_SOFTWARE_TERMS = [
    "software",
    "saas",
    "plattform",
    "platform",
    "cloud",
    "ki",
    "ai",
    "cyber",
    "security",
    "portal",
    "workflow",
    "fachverfahren",
    "dms",
    "ecm",
    "crm",
    "erp",
    "app",
    "digitalisierung",
]

WEAK_OR_SERVICE_ONLY_TERMS = [
    "messe",
    "event",
    "verband",
    "verein",
    "hotel",
    "catering",
    "ticket",
    "agentur fuer veranstaltungen",
    "stadt ",
    "landkreis ",
    "ministerium",
]

CATEGORY_HINTS = {
    "IT": ["ki", "ai", "cloud", "cyber", "security", "daten", "infrastruktur", "software", "plattform"],
    "KOM": ["kommunikation", "intranet", "messenger", "zusammenarbeit", "collaboration"],
    "CROSS": ["projekt", "office", "kollaboration", "workflow", "prozess"],
    "HR": ["personal", "recruiting", "hr", "weiterbildung"],
    "ORG": ["buerger", "bürger", "portal", "fachverfahren", "beschaffung", "vergabe"],
}


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    value = value.replace("ß", "ss")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "anbieter"


def normalize_url(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        return ""
    if not value.startswith(("http://", "https://")):
        value = f"https://{value}"
    return value.rstrip("/")


def host_of(url: str) -> str:
    return urlparse(url).netloc.lower().replace("www.", "")


def host_key(url: str) -> str:
    host = host_of(url)
    parts = host.split(".")
    if len(parts) >= 3 and parts[-2] in {"co", "com", "org"}:
        return ".".join(parts[-3:])
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return host


def clean_label(value: str) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    value = re.sub(r"^(mehr|more|details|weiter|zur webseite|website|partner|aussteller)\s*:?\s*", "", value, flags=re.I)
    return value[:140]


def normalize_company_name(value: str) -> str:
    value = clean_label(value)
    value = re.sub(r"\b(gmbh|ag|se|ug|kg|gbr|mbh|inc|llc|ltd)\b\.?", "", value, flags=re.I)
    value = re.sub(r"[^a-zA-Z0-9äöüÄÖÜß]+", " ", value)
    return re.sub(r"\s+", " ", value).strip().lower()


def noisy_url(url: str) -> bool:
    parsed = urlparse(url)
    host = parsed.netloc.lower().replace("www.", "")
    if not host or any(host == item or host.endswith("." + item) for item in NOISE_HOSTS):
        return True
    path = parsed.path.lower()
    return any(term in path for term in NOISE_PATH_TERMS)


def likely_discovery_page(url: str, label: str) -> bool:
    haystack = f"{url} {label}".lower()
    return any(term in haystack for term in DISCOVERY_PAGE_TERMS)


def likely_profile_page(url: str, label: str) -> bool:
    haystack = f"{url} {label}".lower()
    return any(term in haystack for term in PROFILE_PAGE_TERMS)


def source_hosts(source: dict[str, Any]) -> set[str]:
    hosts = {host_key(source.get("url", ""))}
    for edition in source.get("editions", []):
        hosts.add(host_key(edition.get("edition_url", "")))
    return {host for host in hosts if host}


def read_existing_keys(excel_path: Path, curation_path: Path) -> dict[str, set[str]]:
    names: set[str] = set()
    hosts: set[str] = set()

    if excel_path.exists():
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["Zielkunden"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        for raw in rows[1:]:
            item = dict(zip(headers, raw))
            if item.get("Firma"):
                names.add(slugify(str(item["Firma"])))
            if item.get("Website"):
                hosts.add(host_key(normalize_url(str(item["Website"]))))

    if curation_path.exists():
        curation = json.loads(curation_path.read_text(encoding="utf-8"))
        for entry in curation.get("entries", []):
            action = str(entry.get("crawl_action") or "").strip().lower()
            status = str(entry.get("status") or "").strip().lower()
            if action not in {"crawl", "recrawl", "monitor_existing"} and status not in {"already_in_seed"}:
                continue
            if entry.get("name"):
                names.add(slugify(entry["name"]))
            if entry.get("website"):
                hosts.add(host_key(normalize_url(entry["website"])))

    return {"names": names, "hosts": {host for host in hosts if host}}


async def crawl_page(crawler: AsyncWebCrawler, url: str) -> dict[str, Any]:
    config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        check_robots_txt=True,
        wait_for="body",
        wait_until="domcontentloaded",
        delay_before_return_html=1.0,
        page_timeout=45000,
        remove_overlay_elements=True,
        remove_consent_popups=True,
        scan_full_page=False,
        verbose=False,
    )
    try:
        result = await crawler.arun(url=url, config=config)
        return {
            "url": url,
            "success": bool(result.success),
            "status_code": getattr(result, "status_code", None),
            "html": str(getattr(result, "html", "") or ""),
            "markdown": str(getattr(result, "markdown", "") or ""),
            "error": getattr(result, "error_message", "") or "",
        }
    except Exception as exc:
        return {"url": url, "success": False, "status_code": None, "html": "", "markdown": "", "error": str(exc)}


def score_candidate(candidate: dict[str, Any], source: dict[str, Any]) -> tuple[int, list[str]]:
    label = candidate.get("name", "")
    url = candidate.get("website", "")
    context = f"{label} {url} {candidate.get('context', '')}".lower()
    score = 0
    reasons: list[str] = []

    if candidate.get("link_type") == "external_company":
        score += 2
        reasons.append("externer Anbieterlink")
    if COMPANY_SUFFIX_RE.search(label):
        score += 2
        reasons.append("Firmen-/Software-Signal im Namen")
    if source.get("public_sector_fit") == "high":
        score += 1
        reasons.append("Quelle mit hohem Public-Sector-Fit")
    if any(term in context for term in FIT_TERMS["public_sector"]):
        score += 2
        reasons.append("Public-Sector-Begriff im Kontext")
    if any(term in context for term in FIT_TERMS["software"]):
        score += 2
        reasons.append("Software-/Digital-Signal im Kontext")
    if any(term in context for term in STRONG_SOFTWARE_TERMS):
        score += 1
        reasons.append("starkes Software-Signal")
    if any(term in context for term in WEAK_OR_SERVICE_ONLY_TERMS):
        score -= 2
        reasons.append("moeglicherweise Service-/Event-/Institutionstreffer")
    if host_key(url) in {"bitkom.org", "messe-berlin.de", "messe-muenchen.de", "messe-duesseldorf.de"}:
        score -= 3
        reasons.append("wahrscheinlich Veranstalter-/Verbandseite")
    return score, reasons


def infer_cluster(candidate: dict[str, Any], source: dict[str, Any]) -> str:
    haystack = f"{candidate.get('name', '')} {candidate.get('context', '')} {' '.join(source.get('topic_focus', []))}".lower()
    best = ("CROSS", 0)
    for cluster, terms in CATEGORY_HINTS.items():
        hits = sum(1 for term in terms if term in haystack)
        if hits > best[1]:
            best = (cluster, hits)
    return best[0]


def branch_from_source(source: dict[str, Any]) -> str:
    topics = source.get("topic_focus") or []
    if topics:
        return " / ".join(topics[:3])
    return "Public-Sector-Softwareanbieter"


def extract_candidates(page: dict[str, Any], source: dict[str, Any]) -> list[dict[str, Any]]:
    soup = BeautifulSoup(page.get("html") or "", "html.parser")
    source_host_keys = source_hosts(source)
    found: list[dict[str, Any]] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href") or "").strip()
        if not href or href.startswith(("mailto:", "tel:", "#", "javascript:")):
            continue
        url = normalize_url(urljoin(page["url"] + "/", href).split("#", 1)[0])
        if noisy_url(url):
            continue

        label = clean_label(anchor.get_text(" ", strip=True) or anchor.get("title") or "")
        if len(label) < 3:
            label = host_key(url).split(".")[0].replace("-", " ").title()
        if len(label) < 3:
            continue

        key = host_key(url)
        if not key or key in seen:
            continue
        seen.add(key)

        same_source = key in source_host_keys
        link_type = "source_profile" if same_source else "external_company"
        path = urlparse(url).path.lower()
        if same_source and not any(term in path for term in ["aussteller", "exhibitor", "partner", "sponsor"]):
            continue

        parent_text = clean_label(anchor.parent.get_text(" ", strip=True) if anchor.parent else "")
        candidate = {
            "name": label,
            "slug": slugify(label),
            "website": url,
            "host": key,
            "source_id": source.get("source_id"),
            "source_name": source.get("name"),
            "source_url": page["url"],
            "link_type": link_type,
            "context": parent_text,
        }
        score, reasons = score_candidate(candidate, source)
        candidate["score"] = score
        candidate["score_reasons"] = reasons
        candidate["cluster"] = infer_cluster(candidate, source)
        candidate["branch"] = branch_from_source(source)
        found.append(candidate)

    return found


def discover_source_urls(page: dict[str, Any], source: dict[str, Any], max_urls: int) -> list[str]:
    soup = BeautifulSoup(page.get("html") or "", "html.parser")
    source_host_keys = source_hosts(source)
    scored: dict[str, int] = {}
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href") or "").strip()
        if not href or href.startswith(("mailto:", "tel:", "#", "javascript:")):
            continue
        url = normalize_url(urljoin(page["url"] + "/", href).split("#", 1)[0])
        if noisy_url(url) or host_key(url) not in source_host_keys:
            continue
        label = anchor.get_text(" ", strip=True) or anchor.get("title") or ""
        if not likely_discovery_page(url, label):
            continue
        haystack = f"{url} {label}".lower()
        score = sum(1 for term in DISCOVERY_PAGE_TERMS if term in haystack)
        scored[url] = max(scored.get(url, 0), score)
    return [url for url, _ in sorted(scored.items(), key=lambda item: (-item[1], item[0]))[:max_urls]]


def source_seed_urls(source: dict[str, Any]) -> list[str]:
    urls = [normalize_url(source.get("url", ""))]
    for edition in source.get("editions", []):
        urls.append(normalize_url(edition.get("edition_url", "")))
    seen: set[str] = set()
    result = []
    for url in urls:
        if url and url not in seen:
            result.append(url)
            seen.add(url)
    return result


def entity_key(candidate: dict[str, Any]) -> str:
    host = candidate.get("host") or host_key(candidate.get("website", ""))
    if host:
        return f"host:{host}"
    return f"name:{slugify(normalize_company_name(candidate.get('name', '')))}"


def merge_entities(candidates: list[dict[str, Any]], existing: dict[str, set[str]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        key = entity_key(candidate)
        entity = grouped.setdefault(
            key,
            {
                "entity_key": key,
                "name": candidate.get("name", ""),
                "slug": candidate.get("slug", ""),
                "website": candidate.get("website", ""),
                "host": candidate.get("host", ""),
                "cluster": candidate.get("cluster", "CROSS"),
                "branch": candidate.get("branch", "Public-Sector-Softwareanbieter"),
                "best_score": candidate.get("score", 0),
                "score_reasons": [],
                "sources": [],
                "candidate_variants": [],
            },
        )
        entity["candidate_variants"].append(candidate)
        if candidate.get("score", 0) > entity.get("best_score", 0):
            entity.update(
                {
                    "name": candidate.get("name", entity["name"]),
                    "slug": candidate.get("slug", entity["slug"]),
                    "website": candidate.get("website", entity["website"]),
                    "host": candidate.get("host", entity["host"]),
                    "cluster": candidate.get("cluster", entity["cluster"]),
                    "branch": candidate.get("branch", entity["branch"]),
                    "best_score": candidate.get("score", 0),
                }
            )
        entity["score_reasons"].extend(candidate.get("score_reasons", []))
        entity["sources"].append(
            {
                "source_id": candidate.get("source_id"),
                "source_name": candidate.get("source_name"),
                "source_url": candidate.get("source_url"),
                "context": candidate.get("context", ""),
                "score": candidate.get("score", 0),
            }
        )

    entities = []
    for entity in grouped.values():
        entity["score_reasons"] = sorted(set(entity["score_reasons"]))
        entity["source_count"] = len({source["source_id"] for source in entity["sources"] if source.get("source_id")})
        entity["known_in_excel_or_inbox"] = entity.get("host") in existing["hosts"] or entity.get("slug") in existing["names"]
        entity["prequalification"] = prequalify_entity(entity)
        entities.append(entity)
    return sorted(entities, key=lambda item: (-item["prequalification"]["priority"], -item.get("best_score", 0), item["name"]))


def prequalify_entity(entity: dict[str, Any]) -> dict[str, Any]:
    haystack = " ".join(
        [
            entity.get("name", ""),
            entity.get("website", ""),
            " ".join(entity.get("score_reasons", [])),
            " ".join(source.get("context", "") for source in entity.get("sources", [])),
        ]
    ).lower()
    software_hits = [term for term in STRONG_SOFTWARE_TERMS if term in haystack]
    public_hits = [term for term in FIT_TERMS["public_sector"] if term in haystack]
    service_noise = [term for term in WEAK_OR_SERVICE_ONLY_TERMS if term in haystack]
    score = int(entity.get("best_score", 0)) + len(software_hits) + min(2, entity.get("source_count", 0))
    reasons = []
    if software_hits:
        reasons.append(f"Software-Hinweise: {', '.join(sorted(set(software_hits))[:4])}")
    if public_hits:
        reasons.append(f"Public-Sector-Hinweise: {', '.join(sorted(set(public_hits))[:4])}")
    if entity.get("source_count", 0) > 1:
        reasons.append(f"in {entity['source_count']} Quellen gefunden")
    if entity.get("known_in_excel_or_inbox"):
        reasons.append("bereits in Excel oder Candidate-Inbox bekannt")
    if service_noise:
        reasons.append("moeglicher Nicht-Software-/Institutionstreffer")

    clear_noise = bool(service_noise) and not software_hits
    if entity.get("known_in_excel_or_inbox"):
        bucket = "known_monitor"
        priority = 20
    elif clear_noise:
        bucket = "reject_noise"
        priority = 0
    elif score >= 9 and software_hits and "moeglicher Nicht-Software-/Institutionstreffer" not in reasons:
        bucket = "crawl_next"
        priority = 90
    elif score >= 7 and (software_hits or public_hits):
        bucket = "review_for_crawl"
        priority = 70
    elif score >= 5:
        bucket = "weak_watchlist"
        priority = 40
    else:
        bucket = "reject_noise"
        priority = 0
    return {"bucket": bucket, "priority": priority, "score": score, "reasons": reasons}


def append_to_curation(curation_path: Path, entities: list[dict[str, Any]], existing: dict[str, set[str]], limit: int) -> int:
    payload = json.loads(curation_path.read_text(encoding="utf-8")) if curation_path.exists() else {
        "schema_version": 1,
        "updated_at": "",
        "purpose": "Kuratierte Discovery-Seedliste fuer zusaetzliche Supertools-Kandidaten.",
        "notes": [],
        "entries": [],
    }
    entries = payload.setdefault("entries", [])
    known_hosts = set(existing["hosts"])
    known_names = set(existing["names"])
    for entry in entries:
        if entry.get("website"):
            known_hosts.add(host_key(entry["website"]))
        if entry.get("name"):
            known_names.add(slugify(entry["name"]))

    appended = 0
    start_rank = max([int(entry.get("rank") or 0) for entry in entries] + [2000]) + 1
    for entity in sorted(entities, key=lambda item: (-item["prequalification"]["priority"], -item.get("best_score", 0), item.get("name", ""))):
        if appended >= limit:
            break
        if entity["host"] in known_hosts or entity["slug"] in known_names:
            continue
        if entity["prequalification"]["bucket"] != "crawl_next":
            continue
        entries.append(
            {
                "rank": start_rank + appended,
                "name": entity["name"],
                "submitted_names": sorted({variant.get("name", entity["name"]) for variant in entity["candidate_variants"] if variant.get("name")}),
                "slug": entity["slug"],
                "website": entity["website"],
                "source_bucket": "Broad Multiplier Discovery",
                "status": "discovered_from_multiplier",
                "crawl_action": "crawl",
                "cluster": entity["cluster"],
                "branch": entity["branch"],
                "relevance_score": min(5, max(3, entity["prequalification"]["score"] - 4)),
                "curation_note": "Aus breiter Multiplikatoren-Discovery entdeckt; vor Website-Nutzung redaktionell pruefen.",
                "discovery_sources": sorted({source["source_url"] for source in entity["sources"] if source.get("source_url")}),
                "discovery_score": entity["prequalification"]["score"],
                "discovery_score_reasons": entity["prequalification"]["reasons"],
                "entity_match": {
                    "entity_key": entity["entity_key"],
                    "source_count": entity["source_count"],
                    "best_score": entity["best_score"],
                },
            }
        )
        known_hosts.add(entity["host"])
        known_names.add(entity["slug"])
        appended += 1

    payload["updated_at"] = dt.date.today().isoformat()
    curation_path.parent.mkdir(parents=True, exist_ok=True)
    curation_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return appended


def report(
    run_meta: dict[str, Any],
    candidates: list[dict[str, Any]],
    entities: list[dict[str, Any]],
    pages: list[dict[str, Any]],
) -> str:
    buckets = {}
    for entity in entities:
        bucket = entity["prequalification"]["bucket"]
        buckets[bucket] = buckets.get(bucket, 0) + 1
    lines = [
        "# Multiplikatoren-Discovery-Run",
        "",
        f"Run: `{run_meta['run_id']}`",
        f"Zeit: {run_meta['started_at']}",
        f"Quellen im Lauf: {run_meta['source_count']}",
        f"Seiten erfolgreich: {sum(1 for page in pages if page.get('success'))}/{len(pages)}",
        f"Kandidaten brutto: {len(candidates)}",
        f"Entitaeten nach Matching: {len(entities)}",
        f"Vorqualifiziert `crawl_next`: {buckets.get('crawl_next', 0)}",
        f"Vorqualifiziert `review_for_crawl`: {buckets.get('review_for_crawl', 0)}",
        f"Bekannt / Monitoring: {buckets.get('known_monitor', 0)}",
        f"In Candidate-Inbox geschrieben: {run_meta.get('appended_to_curation', 0)}",
        "",
        "## Vorqualifizierte Entities",
        "",
        "| Bucket | Score | Anbieter | Quellen | Link | Gründe |",
        "| --- | ---: | --- | ---: | --- | --- |",
    ]
    for item in entities[:100]:
        reasons = "; ".join(item["prequalification"].get("reasons", []))
        lines.append(
            f"| `{item['prequalification']['bucket']}` | {item['prequalification']['score']} | {item['name'].replace('|', '/')} | {item.get('source_count', 0)} | {item['website']} | {reasons.replace('|', '/')} |"
        )

    lines.extend(["", "## Top-Rohkandidaten", "", "| Score | Anbieter | Quelle | Link | Gründe |", "| ---: | --- | --- | --- | --- |"])
    for item in sorted(candidates, key=lambda child: (-child.get("score", 0), child.get("name", "")))[:80]:
        reasons = ", ".join(item.get("score_reasons", []))
        lines.append(
            f"| {item.get('score', 0)} | {item['name'].replace('|', '/')} | {item['source_name']} | {item['website']} | {reasons.replace('|', '/')} |"
        )

    lines.extend(["", "## Crawl-Probleme", ""])
    failed = [page for page in pages if not page.get("success")]
    if failed:
        for page in failed:
            lines.append(f"- {page['url']} - {page.get('error') or page.get('status_code')}")
    else:
        lines.append("- Keine harten Crawling-Fehler.")
    return "\n".join(lines) + "\n"


async def run(args: argparse.Namespace) -> None:
    started = dt.datetime.now(dt.timezone.utc).astimezone()
    run_id = args.run_id or f"multiplier-discovery-{started.strftime('%Y%m%d-%H%M%S')}"
    output_dir = Path(args.out) / run_id
    output_dir.mkdir(parents=True, exist_ok=True)

    multipliers = json.loads(Path(args.multipliers).read_text(encoding="utf-8"))
    sources = multipliers.get("entries", [])
    if args.fit:
        sources = [source for source in sources if source.get("public_sector_fit") == args.fit]
    sources = sources[args.offset : args.offset + args.limit]
    existing = read_existing_keys(Path(args.excel), Path(args.curation))

    browser_config = BrowserConfig(headless=True, viewport_width=1440, viewport_height=1000, text_mode=False, verbose=False)
    pages: list[dict[str, Any]] = []
    raw_candidates: list[dict[str, Any]] = []
    candidates_by_host: dict[str, dict[str, Any]] = {}

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for index, source in enumerate(sources, start=1):
            seed_urls = source_seed_urls(source)
            if not seed_urls:
                continue
            print(f"[{index}/{len(sources)}] Broad discovering {source['name']} ({len(seed_urls)} seed URLs)")
            source_pages: list[dict[str, Any]] = []
            seen_urls: set[str] = set()
            queue = list(seed_urls)

            while queue and len(source_pages) < args.pages_per_source:
                url = queue.pop(0)
                if url in seen_urls:
                    continue
                seen_urls.add(url)
                page = await crawl_page(crawler, url)
                pages.append(page)
                source_pages.append(page)
                if not page.get("success"):
                    continue
                for discovered_url in discover_source_urls(page, source, args.discovery_links_per_source):
                    if discovered_url not in seen_urls and discovered_url not in queue:
                        queue.append(discovered_url)

            profile_queue: list[str] = []
            for page in source_pages:
                if not page.get("success"):
                    continue
                page_candidates = extract_candidates(page, source)
                raw_candidates.extend(page_candidates)
                for candidate in page_candidates:
                    if candidate["link_type"] == "source_profile" and likely_profile_page(candidate["website"], candidate["name"]):
                        profile_queue.append(candidate["website"])
                    if candidate["link_type"] != "external_company":
                        continue
                    key = candidate["host"]
                    previous = candidates_by_host.get(key)
                    if previous is None or candidate.get("score", 0) > previous.get("score", 0):
                        candidates_by_host[key] = candidate

            for profile_url in profile_queue[: args.profile_pages_per_source]:
                if profile_url in seen_urls:
                    continue
                seen_urls.add(profile_url)
                page = await crawl_page(crawler, profile_url)
                pages.append(page)
                if not page.get("success"):
                    continue
                page_candidates = extract_candidates(page, source)
                raw_candidates.extend(page_candidates)
                for candidate in page_candidates:
                    if candidate["link_type"] != "external_company":
                        continue
                key = candidate["host"]
                previous = candidates_by_host.get(key)
                if previous is None or candidate.get("score", 0) > previous.get("score", 0):
                    candidates_by_host[key] = candidate

    candidates = sorted(candidates_by_host.values(), key=lambda item: (-item.get("score", 0), item["name"]))
    entities = merge_entities(candidates, existing)
    prequalified = [entity for entity in entities if entity["prequalification"]["bucket"] in {"crawl_next", "review_for_crawl"}]
    appended = append_to_curation(Path(args.curation), entities, existing, args.append_limit) if args.append else 0

    run_meta = {
        "run_id": run_id,
        "started_at": started.isoformat(timespec="seconds"),
        "source_count": len(sources),
        "page_count": len(pages),
        "raw_candidate_count": len(raw_candidates),
        "candidate_count": len(candidates),
        "entity_count": len(entities),
        "prequalified_count": len(prequalified),
        "prequalified_crawl_next": sum(1 for item in entities if item["prequalification"]["bucket"] == "crawl_next"),
        "prequalified_review_for_crawl": sum(1 for item in entities if item["prequalification"]["bucket"] == "review_for_crawl"),
        "appended_to_curation": appended,
        "multipliers": args.multipliers,
        "curation": args.curation,
        "pages_per_source": args.pages_per_source,
        "discovery_links_per_source": args.discovery_links_per_source,
        "profile_pages_per_source": args.profile_pages_per_source,
    }
    (output_dir / "run-meta.json").write_text(json.dumps(run_meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "raw-link-candidates.json").write_text(json.dumps(raw_candidates, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "multiplier-candidates.json").write_text(json.dumps(candidates, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "entity-matches.json").write_text(json.dumps(entities, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "prequalified-candidates.json").write_text(json.dumps(prequalified, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "pages.json").write_text(json.dumps(pages, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (output_dir / "review-report.md").write_text(report(run_meta, candidates, entities, pages), encoding="utf-8")
    print("")
    print(f"Wrote {output_dir / 'review-report.md'}")
    print(f"Wrote {output_dir / 'entity-matches.json'}")
    print(f"Wrote {output_dir / 'prequalified-candidates.json'}")
    print(f"Wrote {output_dir / 'multiplier-candidates.json'}")
    print(f"Appended to curation: {appended}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--multipliers", default=DEFAULT_MULTIPLIERS)
    parser.add_argument("--curation", default=DEFAULT_CURATION)
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--out", default=DEFAULT_OUT)
    parser.add_argument("--run-id")
    parser.add_argument("--limit", type=int, default=30)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--fit", choices=["high", "medium", "low", "unknown"])
    parser.add_argument("--pages-per-source", type=int, default=4)
    parser.add_argument("--discovery-links-per-source", type=int, default=8)
    parser.add_argument("--profile-pages-per-source", type=int, default=10)
    parser.add_argument("--append", action="store_true")
    parser.add_argument("--append-limit", type=int, default=20)
    return parser.parse_args()


if __name__ == "__main__":
    asyncio.run(run(parse_args()))
