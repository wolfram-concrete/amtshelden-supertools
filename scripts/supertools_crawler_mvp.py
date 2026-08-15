#!/usr/bin/env python3
"""Weekly Supertools crawler MVP.

Reads the partner Excel seed list, crawls public provider pages with Crawl4AI,
extracts editorial review signals, and writes a human-in-the-loop report.
"""

from __future__ import annotations

import argparse
import asyncio
import contextlib
import dataclasses
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from crawl4ai import AsyncWebCrawler, BrowserConfig, CacheMode, CrawlerRunConfig
from openpyxl import load_workbook


DEFAULT_EXCEL = "Amtshelden_Zielkundenliste_Sponsoring_2026 (1).xlsx"
DEFAULT_CURATION_SEEDS = "data/crawler/discovery/curation-seeds.json"
DEFAULT_STATE = "data/crawler/state/products.json"
DEFAULT_RUNS = "data/crawler/runs"

PRIORITY_LINK_TERMS = [
    "datenschutz",
    "privacy",
    "impressum",
    "security",
    "sicherheit",
    "compliance",
    "dsgvo",
    "gdpr",
    "referenz",
    "referenzen",
    "kunden",
    "case",
    "cases",
    "public",
    "verwaltung",
    "behoerde",
    "behorde",
    "kommun",
    "produkt",
    "loesung",
    "lösung",
]

PRODUCT_IMAGE_POSITIVE_TERMS = {
    "dashboard": 10,
    "screenshot": 10,
    "software": 7,
    "platform": 7,
    "plattform": 7,
    "interface": 7,
    "workspace": 6,
    "workflow": 6,
    "cockpit": 6,
    "portal": 5,
    "product": 4,
    "produkt": 4,
    "app": 4,
    "ui": 4,
}

PRODUCT_IMAGE_NEGATIVE_TERMS = {
    "logo": -12,
    "icon": -10,
    "favicon": -12,
    "avatar": -10,
    "portrait": -9,
    "team": -8,
    "people": -7,
    "person": -7,
    "webinar": -7,
    "event": -6,
    "award": -8,
    "badge": -8,
    "partner": -7,
    "customer": -6,
    "kunde": -6,
    "social": -8,
    "instagram": -10,
    "linkedin": -10,
}

SIGNAL_TERMS = {
    "public_sector": [
        "fuer behoerden",
        "für behörden",
        "behoerden und verwaltungen",
        "behörden und verwaltungen",
        "oeffentliche verwaltung",
        "öffentliche verwaltung",
        "oeffentlichen verwaltung",
        "öffentlichen verwaltung",
        "kommunen",
        "kommunalverwaltung",
        "landkreis",
        "ministerium",
        "oeffentlicher dienst",
        "öffentlicher dienst",
        "oeffentlicher sektor",
        "öffentlicher sektor",
        "egovernment",
        "e-government",
        "ozg",
    ],
    "privacy": [
        "dsgvo",
        "gdpr",
        "datenschutz",
        "auftragsverarbeitung",
        "avv",
        "tom",
        "technische und organisatorische massnahmen",
        "technische und organisatorische maßnahmen",
    ],
    "hosting": [
        "serverstandort",
        "rechenzentrum",
        "all data is stored securely in germany",
        "data is stored securely in germany",
        "stored securely in germany",
        "hosting in deutschland",
        "in deutschland gehostet",
        "daten in deutschland",
        "deutsches rechenzentrum",
        "deutschem rechenzentrum",
        "standort deutschland",
        "deutsche cloud",
        "hosting in der eu",
        "daten in der eu",
        "eu-rechenzentrum",
        "eu rechenzentrum",
        "europaeische union",
        "europäische union",
        "hosting",
        "souveraene cloud",
        "souveräne cloud",
    ],
    "security": [
        "iso 27001",
        "iso/iec 27001",
        "bsi c5",
        "c5 testat",
        "c5-typ",
        "bsi",
        "tisax",
        "it-grundschutz",
        "verschluesselung",
        "verschlüsselung",
        "zero trust",
    ],
    "accessibility": [
        "barrierefrei",
        "barrierefreiheit",
        "bitv",
        "bfsg",
        "wcag",
        "inklusion",
        "accessibility",
    ],
    "operation": [
        "saas",
        "cloud-software",
        "cloud software",
        "cloud-plattform",
        "cloud plattform",
        "cloud-lösung",
        "cloud lösung",
        "cloud solution",
        "on-premise",
        "on premise",
        "onpremise",
        "hybrid",
        "private cloud",
    ],
    "references": [
        "referenz",
        "referenzen",
        "kunden",
        "kundenlogo",
        "kundenstimme",
        "case study",
        "fallstudie",
        "erfolgsgeschichte",
        "stadt ",
        "landkreis ",
        "ministerium",
    ],
}

SIGNAL_LABELS = {
    "public_sector": "Behoerden-/Verwaltungsbezug",
    "privacy": "DSGVO/Datenschutz",
    "hosting": "Hosting/Serverstandort",
    "security": "Sicherheit/Zertifizierung",
    "accessibility": "Barrierefreiheit",
    "operation": "Betriebsmodell",
    "references": "Referenzen/Cases",
}

AVAILABILITY_SCOPE_LABELS = {
    "nationwide": "bundesweit",
    "federal_state": "bundeslandspezifisch",
    "regional": "regional",
    "unknown": "unklar",
}

FEDERAL_STATE_ALIASES = {
    "Baden-Wuerttemberg": ["baden-wuerttemberg", "baden-württemberg"],
    "Bayern": ["bayern", "bavaria"],
    "Berlin": ["berlin"],
    "Brandenburg": ["brandenburg"],
    "Bremen": ["bremen"],
    "Hamburg": ["hamburg"],
    "Hessen": ["hessen"],
    "Mecklenburg-Vorpommern": ["mecklenburg-vorpommern", "mecklenburg vorpommern"],
    "Niedersachsen": ["niedersachsen"],
    "Nordrhein-Westfalen": ["nordrhein-westfalen", "nordrhein westfalen", "nrw"],
    "Rheinland-Pfalz": ["rheinland-pfalz", "rheinland pfalz"],
    "Saarland": ["saarland"],
    "Sachsen": ["sachsen"],
    "Sachsen-Anhalt": ["sachsen-anhalt", "sachsen anhalt"],
    "Schleswig-Holstein": ["schleswig-holstein", "schleswig holstein"],
    "Thueringen": ["thueringen", "thüringen"],
}

NATIONWIDE_TERMS = [
    "bundesweit",
    "deutschlandweit",
    "in ganz deutschland",
    "deutschlandweit taetig",
    "deutschlandweit tätig",
    "bundesweit taetig",
    "bundesweit tätig",
    "fuer alle bundeslaender",
    "für alle bundesländer",
    "alle bundeslaender",
    "alle bundesländer",
]

REGIONAL_TERMS = [
    "regional",
    "regionale",
    "regionaler",
    "regionale verwaltung",
    "vor ort",
    "im umkreis",
    "in der region",
]

STATE_CONTEXT_TERMS = [
    "kommunen",
    "kommune",
    "behoerden",
    "behörden",
    "verwaltung",
    "verwaltungen",
    "landesverwaltung",
    "landesloesung",
    "landeslösung",
    "oeffentlicher sektor",
    "öffentlicher sektor",
]

STATE_RESTRICTION_TERMS = [
    "nur",
    "ausschliesslich",
    "ausschließlich",
    "derzeit",
    "aktuell",
    "verfuegbar",
    "verfügbar",
    "begleiten",
    "betreuen",
]

MISSING_RULES = {
    "privacy": "Keine belastbare DSGVO-/Datenschutz-Aussage gefunden.",
    "hosting": "Kein Serverstandort oder Hosting-Ort oeffentlich auffindbar.",
    "operation": "Kein klares Betriebsmodell (Cloud/On-Premise/Hybrid) gefunden.",
    "public_sector": "Keine klare Behoerden- oder Verwaltungsreferenz gefunden.",
}

CONTENT_KEYWORDS = {
    "use_case": ["use-case", "use case", "anwendungsfall", "loesung/", "lösung/"],
    "case_study": [
        "case study",
        "case-study",
        "case",
        "fallstudie",
        "erfolgsgeschichte",
        "referenzbericht",
        "customer story",
        "kundengeschichte",
    ],
    "webinar": ["webinar", "webinare", "seminar", "online-seminar", "event", "events"],
    "whitepaper": ["whitepaper", "white paper", "ebook", "e-book", "guide", "leitfaden"],
    "blog_article": ["blog", "magazin", "artikel", "insights", "wissen", "news"],
    "video": ["video", "vimeo"],
    "download": [".pdf", "download", "broschuere", "broschüre", "factsheet", "datenblatt"],
}

NOISY_CONTENT_SOURCE_TERMS = [
    "datenschutz",
    "privacy",
    "data-protection",
    "impressum",
    "legal",
    "terms",
    "cookie",
    "newsletter",
]

NOISY_CONTENT_DOMAINS = {
    "facebook.com",
    "de-de.facebook.com",
    "linkedin.com",
    "instagram.com",
    "twitter.com",
    "x.com",
    "developers.facebook.com",
}


@dataclasses.dataclass
class Seed:
    rank: int | None
    company: str
    website: str
    branch: str
    cluster: str
    relevance_score: int | None
    city: str
    employees: str
    phone: str
    email: str
    note: str
    source: str = "excel"


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    value = value.replace("ß", "ss")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "anbieter"


def normalize_url(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        return value
    if not value.startswith(("http://", "https://")):
        value = f"https://{value}"
    return value.rstrip("/")


def read_seeds(path: Path, cluster: str | None, score_min: int | None) -> list[Seed]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Zielkunden"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    seeds: list[Seed] = []

    for raw in rows[1:]:
        item = dict(zip(headers, raw))
        if not item.get("Firma") or not item.get("Website"):
            continue
        item_cluster = str(item.get("Topic-Cluster") or "").strip()
        item_score = item.get("Relevanz-Score (1-5)")
        if cluster and item_cluster.upper() != cluster.upper():
            continue
        if score_min and (not item_score or int(item_score) < score_min):
            continue
        seeds.append(
            Seed(
                rank=int(item["Rang"]) if item.get("Rang") else None,
                company=str(item.get("Firma") or "").strip(),
                website=normalize_url(str(item.get("Website") or "")),
                branch=str(item.get("Branche") or "").strip(),
                cluster=item_cluster,
                relevance_score=int(item_score) if item_score else None,
                city=str(item.get("Hauptsitz") or "").strip(),
                employees=str(item.get("Mitarbeiter") or "").strip(),
                phone=str(item.get("Telefon (Zentrale)") or "").strip(),
                email=str(item.get("E-Mail (Unternehmen)") or "").strip(),
                note=str(item.get("Begründung / Notiz") or "").strip(),
                source="excel",
            )
        )
    return seeds


def read_curation_seeds(path: Path, cluster: str | None, score_min: int | None) -> list[Seed]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    seeds: list[Seed] = []
    for entry in payload.get("entries", []):
        action = str(entry.get("crawl_action") or "").strip().lower()
        if action not in {"crawl", "recrawl"}:
            continue
        item_cluster = str(entry.get("cluster") or "").strip()
        item_score = entry.get("relevance_score")
        if cluster and item_cluster.upper() != cluster.upper():
            continue
        if score_min and (not item_score or int(item_score) < score_min):
            continue
        website = normalize_url(str(entry.get("website") or ""))
        if not entry.get("name") or not website:
            continue
        note_parts = [
            str(entry.get("curation_note") or "").strip(),
            f"Quelle: {entry.get('source_bucket')}" if entry.get("source_bucket") else "",
            f"Status: {entry.get('status')}" if entry.get("status") else "",
        ]
        seeds.append(
            Seed(
                rank=int(entry["rank"]) if entry.get("rank") else None,
                company=str(entry.get("name") or "").strip(),
                website=website,
                branch=str(entry.get("branch") or "").strip(),
                cluster=item_cluster,
                relevance_score=int(item_score) if item_score else None,
                city=str(entry.get("city") or "").strip(),
                employees=str(entry.get("employees") or "").strip(),
                phone=str(entry.get("phone") or "").strip(),
                email=str(entry.get("email") or "").strip(),
                note=" | ".join(part for part in note_parts if part),
                source="curation",
            )
        )
    return seeds


def merge_seeds(excel_seeds: list[Seed], curation_seeds: list[Seed]) -> list[Seed]:
    merged: list[Seed] = []
    seen_keys: set[str] = set()

    for seed in [*excel_seeds, *curation_seeds]:
        host = urlparse(seed.website).netloc.replace("www.", "").lower()
        keys = {slugify(seed.company), host}
        if seen_keys & keys:
            continue
        merged.append(seed)
        seen_keys.update(key for key in keys if key)
    return merged


def load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"products": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def same_host(url: str, candidate: str) -> bool:
    base = urlparse(url).netloc.replace("www.", "")
    host = urlparse(candidate).netloc.replace("www.", "")
    return bool(host) and (host == base or host.endswith("." + base))


def link_priority(link: str, text: str) -> int:
    haystack = f"{link} {text}".lower()
    for index, term in enumerate(PRIORITY_LINK_TERMS):
        if term in haystack:
            return len(PRIORITY_LINK_TERMS) - index
    return 0


def discover_priority_links(base_url: str, html: str, max_pages: int) -> list[str]:
    soup = BeautifulSoup(html or "", "html.parser")
    scored: dict[str, int] = {}
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href") or "").strip()
        if not href or href.startswith(("mailto:", "tel:", "#", "javascript:")):
            continue
        absolute = urljoin(base_url + "/", href).split("#", 1)[0].rstrip("/")
        if not same_host(base_url, absolute):
            continue
        priority = link_priority(absolute, anchor.get_text(" ", strip=True))
        if priority:
            scored[absolute] = max(scored.get(absolute, 0), priority)

    fallback_paths = [
        "/datenschutz",
        "/impressum",
        "/security",
        "/sicherheit",
        "/compliance",
        "/referenzen",
        "/kunden",
    ]
    for path in fallback_paths:
        absolute = urljoin(base_url + "/", path.lstrip("/"))
        scored.setdefault(absolute.rstrip("/"), link_priority(absolute, ""))

    ordered = sorted(scored.items(), key=lambda item: (-item[1], item[0]))
    return [url for url, _ in ordered[: max(0, max_pages - 1)]]


async def crawl_url(crawler: AsyncWebCrawler, url: str) -> dict[str, Any]:
    config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        check_robots_txt=True,
        wait_for="body",
        wait_until="domcontentloaded",
        delay_before_return_html=1.0,
        page_timeout=45000,
        remove_overlay_elements=True,
        remove_consent_popups=True,
        scan_full_page=False,
        verbose=False,
    )
    try:
        result = await crawler.arun(url=url, config=config)
        return {
            "url": url,
            "success": bool(result.success),
            "status_code": getattr(result, "status_code", None),
            "ok_status": bool(result.success) and (getattr(result, "status_code", 0) or 0) < 400,
            "error": getattr(result, "error_message", "") or "",
            "markdown": str(getattr(result, "markdown", "") or ""),
            "html": str(getattr(result, "html", "") or ""),
        }
    except Exception as exc:
        return {
            "url": url,
            "success": False,
            "status_code": None,
            "ok_status": False,
            "error": str(exc),
            "markdown": "",
            "html": "",
        }


def clean_text(value: str, max_chars: int = 180000) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    return value[:max_chars]


def find_snippets(text: str, terms: list[str], limit: int = 5) -> list[str]:
    snippets: list[str] = []
    seen: set[str] = set()
    for term in terms:
        pattern = re.escape(term)
        if term == "kunden":
            pattern = r"(?<![a-zäöüß])kunden(?![a-zäöüß])"
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        start = match.start()
        left = max(0, start - 140)
        right = min(len(text), start + 260)
        snippet = text[left:right].strip()
        snippet = re.sub(r"\s+", " ", snippet)
        key = snippet[:100].lower()
        if key not in seen:
            snippets.append(snippet)
            seen.add(key)
        if len(snippets) >= limit:
            break
    return snippets


def find_snippet_details(page: dict[str, Any], terms: list[str], limit: int = 3) -> list[dict[str, str]]:
    text = clean_text(page.get("markdown") or "")
    details: list[dict[str, str]] = []
    seen: set[str] = set()
    for term in terms:
        pattern = re.escape(term)
        if term == "kunden":
            pattern = r"(?<![a-zäöüß])kunden(?![a-zäöüß])"
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        start = match.start()
        left = max(0, start - 120)
        right = min(len(text), start + 260)
        snippet = re.sub(r"\s+", " ", text[left:right].strip())
        key = f"{page['url']}:{snippet[:100].lower()}"
        if key not in seen:
            details.append({"url": page["url"], "term": term, "snippet": snippet})
            seen.add(key)
        if len(details) >= limit:
            break
    return details


def infer_operation(text: str) -> list[str]:
    found = []
    checks = {
        "Cloud/SaaS": [
            "saas",
            "cloud-software",
            "cloud software",
            "cloud-plattform",
            "cloud plattform",
            "cloud-lösung",
            "cloud lösung",
            "cloud solution",
        ],
        "On-Premise": ["on-premise", "on premise", "onpremise"],
        "Hybrid": ["hybrid"],
        "Private Cloud": ["private cloud"],
    }
    lower = text.lower()
    for label, terms in checks.items():
        if any(term in lower for term in terms):
            found.append(label)
    return found


def unique_preserve(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result = []
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result


def snippet_at(text: str, start: int, left_chars: int = 150, right_chars: int = 300) -> str:
    left = max(0, start - left_chars)
    right = min(len(text), start + right_chars)
    return re.sub(r"\s+", " ", text[left:right].strip())


def availability_term_details(
    page: dict[str, Any],
    terms: list[str],
    limit: int = 3,
) -> list[dict[str, str]]:
    text = clean_text(page.get("markdown") or "")
    details = []
    seen: set[str] = set()
    for term in terms:
        match = re.search(re.escape(term), text, flags=re.IGNORECASE)
        if not match:
            continue
        snippet = snippet_at(text, match.start())
        key = f"{page['url']}:{snippet[:120].lower()}"
        if key in seen:
            continue
        details.append({"url": page["url"], "term": term, "snippet": snippet})
        seen.add(key)
        if len(details) >= limit:
            break
    return details


def state_details(page: dict[str, Any], limit: int = 4) -> list[dict[str, str]]:
    text = clean_text(page.get("markdown") or "")
    lower = text.lower()
    details = []
    seen: set[str] = set()
    for state, aliases in FEDERAL_STATE_ALIASES.items():
        for alias in aliases:
            for match in re.finditer(re.escape(alias), lower, flags=re.IGNORECASE):
                snippet = snippet_at(text, match.start())
                snippet_lower = snippet.lower()
                has_context = any(term in snippet_lower for term in STATE_CONTEXT_TERMS)
                has_restriction = any(term in snippet_lower for term in STATE_RESTRICTION_TERMS)
                has_state_offer_pattern = bool(
                    re.search(
                        rf"(fuer|für)\s+(kommunen|behoerden|behörden|verwaltungen?).{{0,80}}\b{re.escape(alias)}\b",
                        snippet_lower,
                    )
                )
                if not (has_state_offer_pattern or (has_context and has_restriction)):
                    continue
                key = f"{page['url']}:{state}:{snippet[:120].lower()}"
                if key in seen:
                    continue
                details.append({"url": page["url"], "term": alias, "region": state, "snippet": snippet})
                seen.add(key)
                if len(details) >= limit:
                    return details
    return details


def infer_availability(pages: list[dict[str, Any]]) -> dict[str, Any]:
    ok_pages = [page for page in pages if page.get("ok_status")]
    nationwide: list[dict[str, str]] = []
    state_based: list[dict[str, str]] = []
    regional: list[dict[str, str]] = []

    for page in ok_pages:
        nationwide.extend(availability_term_details(page, NATIONWIDE_TERMS, limit=2))
        state_based.extend(state_details(page, limit=3))
        regional.extend(availability_term_details(page, REGIONAL_TERMS, limit=2))

    if nationwide:
        return {
            "scope": "nationwide",
            "label": AVAILABILITY_SCOPE_LABELS["nationwide"],
            "regions": [],
            "confidence": "hoch",
            "needs_review": False,
            "evidence": nationwide[:4],
        }

    if state_based:
        regions = unique_preserve([entry["region"] for entry in state_based if entry.get("region")])
        return {
            "scope": "federal_state",
            "label": AVAILABILITY_SCOPE_LABELS["federal_state"],
            "regions": regions,
            "confidence": "mittel",
            "needs_review": True,
            "evidence": state_based[:4],
        }

    if regional:
        return {
            "scope": "regional",
            "label": AVAILABILITY_SCOPE_LABELS["regional"],
            "regions": [],
            "confidence": "niedrig",
            "needs_review": True,
            "evidence": regional[:4],
        }

    return {
        "scope": "unknown",
        "label": AVAILABILITY_SCOPE_LABELS["unknown"],
        "regions": [],
        "confidence": "offen",
        "needs_review": True,
        "evidence": [],
    }


def youtube_id(url: str) -> str | None:
    parsed = urlparse(url)
    host = parsed.netloc.replace("www.", "")
    if host == "youtu.be":
        candidate = parsed.path.strip("/").split("/")[0]
        return candidate or None
    if host in {"youtube.com", "m.youtube.com"}:
        if parsed.path == "/watch":
            query = dict(part.split("=", 1) for part in parsed.query.split("&") if "=" in part)
            return query.get("v")
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 2 and parts[0] in {"embed", "shorts", "live"}:
            return parts[1]
    if "youtube-nocookie.com" in host:
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 2 and parts[0] == "embed":
            return parts[1]
    return None


def classify_content_piece(url: str, label: str) -> str | None:
    haystack = f"{url} {label}".lower()
    if youtube_id(url):
        return "youtube"
    if "youtube.com" in haystack or "youtu.be" in haystack:
        return None
    if "vimeo.com" in haystack and re.search(r"vimeo\.com/(?:video/)?\d+", haystack):
        return "video"
    for kind, keywords in CONTENT_KEYWORDS.items():
        if any(keyword in haystack for keyword in keywords):
            return kind
    return None


def normalize_link(base_url: str, href: str) -> str | None:
    href = str(href or "").strip()
    if not href or href.startswith(("mailto:", "tel:", "#", "javascript:")):
        return None
    return urljoin(base_url + "/", href).split("#", 1)[0]


def content_piece_key(piece: dict[str, Any]) -> str:
    if piece.get("video_id"):
        return f"{piece.get('kind')}:{piece.get('video_id')}"
    return f"{piece.get('kind')}:{piece.get('url')}"


def clean_content_title(title: str, url: str) -> str:
    title = truncate(title, 140)
    if len(title) >= 3 and not re.fullmatch(r"[\W_]+", title):
        return title
    filename = Path(urlparse(url).path).name
    if filename:
        return truncate(filename.replace("-", " ").replace("_", " "), 140)
    return url


def noisy_content_source(url: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    return any(term in path for term in NOISY_CONTENT_SOURCE_TERMS)


def noisy_content_target(url: str) -> bool:
    parsed = urlparse(url)
    host = parsed.netloc.replace("www.", "").lower()
    if host in NOISY_CONTENT_DOMAINS:
        return True
    path = parsed.path.lower()
    if any(term in path for term in ["privacy", "datenschutz", "impressum", "legal", "terms", "cookie"]):
        return True
    return False


def extract_content_pieces(pages: list[dict[str, Any]], limit: int = 20) -> list[dict[str, Any]]:
    pieces: list[dict[str, Any]] = []
    seen: set[str] = set()

    for page in pages:
        if not page.get("ok_status"):
            continue
        source_url = page["url"]
        if noisy_content_source(source_url):
            continue
        soup = BeautifulSoup(page.get("html") or "", "html.parser")

        for frame in soup.find_all(["iframe", "embed"], src=True):
            url = normalize_link(source_url, frame.get("src"))
            if not url:
                continue
            if noisy_content_target(url):
                continue
            kind = classify_content_piece(url, frame.get("title") or "")
            if not kind:
                continue
            video_id = youtube_id(url)
            piece = {
                "kind": kind,
                "title": clean_content_title(frame.get("title") or "Eingebettetes Video", url),
                "url": url,
                "source_url": source_url,
                "platform": "youtube" if video_id else "embed",
                "video_id": video_id,
                "thumbnail_url": f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg" if video_id else None,
            }
            key = content_piece_key(piece)
            if key not in seen:
                pieces.append(piece)
                seen.add(key)

        for anchor in soup.find_all("a", href=True):
            url = normalize_link(source_url, anchor.get("href"))
            if not url:
                continue
            if noisy_content_target(url):
                continue
            title = anchor.get_text(" ", strip=True) or anchor.get("title") or url
            kind = classify_content_piece(url, title)
            if not kind:
                continue
            video_id = youtube_id(url)
            piece = {
                "kind": kind,
                "title": clean_content_title(title, url),
                "url": url,
                "source_url": source_url,
                "platform": "youtube" if video_id else urlparse(url).netloc.replace("www.", ""),
                "video_id": video_id,
                "thumbnail_url": f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg" if video_id else None,
            }
            key = content_piece_key(piece)
            if key not in seen:
                pieces.append(piece)
                seen.add(key)
            if len(pieces) >= limit:
                return pieces

    return pieces[:limit]


def largest_srcset_url(source_url: str, srcset: str) -> str | None:
    options: list[tuple[int, str]] = []
    for entry in (srcset or "").split(","):
        parts = entry.strip().split()
        if not parts:
            continue
        width = 0
        if len(parts) > 1 and parts[-1].endswith("w"):
            with contextlib.suppress(ValueError):
                width = int(parts[-1][:-1])
        normalized = normalize_link(source_url, parts[0])
        if normalized:
            options.append((width, normalized))
    return max(options, default=(0, ""))[1] or None


def product_image_score(url: str, label: str, source_url: str) -> tuple[int, list[str]]:
    haystack = f"{url} {label}".lower()
    score = 0
    reasons: list[str] = []
    for term, weight in PRODUCT_IMAGE_POSITIVE_TERMS.items():
        if re.search(rf"(?<![a-z0-9]){re.escape(term)}(?![a-z0-9])", haystack):
            score += weight
            reasons.append(f"+{weight}:{term}")
    for term, weight in PRODUCT_IMAGE_NEGATIVE_TERMS.items():
        if term in haystack:
            score += weight
            reasons.append(f"{weight}:{term}")

    dimension_match = re.search(r"(?<!\d)(\d{3,5})x(\d{3,5})(?!\d)", url)
    if dimension_match:
        width, height = map(int, dimension_match.groups())
        if width >= 900 and height >= 450:
            score += 5
            reasons.append("+5:grosse-quelldatei")
        elif width <= 320 or height <= 240:
            score -= 8
            reasons.append("-8:kleine-quelldatei")

    if any(term in source_url.lower() for term in ["produkt", "product", "plattform", "platform", "feature"]):
        score += 3
        reasons.append("+3:produktseite")
    return score, reasons


def extract_product_image_candidates(
    pages: list[dict[str, Any]],
    limit: int = 30,
) -> list[dict[str, Any]]:
    candidates: dict[str, dict[str, Any]] = {}
    for page in pages:
        if not page.get("ok_status") or noisy_content_source(page["url"]):
            continue
        source_url = page["url"]
        soup = BeautifulSoup(page.get("html") or "", "html.parser")
        for image in soup.find_all("img"):
            candidate_urls: list[str] = []
            for attribute in ["src", "data-src", "data-lazy-src", "data-original"]:
                normalized = normalize_link(source_url, image.get(attribute))
                if normalized:
                    candidate_urls.append(normalized)
            for attribute in ["srcset", "data-srcset"]:
                largest = largest_srcset_url(source_url, image.get(attribute) or "")
                if largest:
                    candidate_urls.append(largest)

            label = " ".join(
                part
                for part in [
                    image.get("alt"),
                    image.get("title"),
                    image.get("aria-label"),
                ]
                if part
            ).strip()
            for image_url in candidate_urls:
                parsed = urlparse(image_url)
                if parsed.scheme not in {"http", "https"}:
                    continue
                if parsed.path.lower().endswith((".svg", ".gif")):
                    continue
                score, reasons = product_image_score(image_url, label, source_url)
                record = {
                    "url": image_url,
                    "source_url": source_url,
                    "alt": label,
                    "score": score,
                    "reasons": reasons,
                    "review_status": "needs_review",
                }
                previous = candidates.get(image_url)
                if previous is None or score > previous["score"]:
                    candidates[image_url] = record

    return sorted(
        candidates.values(),
        key=lambda candidate: (-candidate["score"], candidate["url"]),
    )[:limit]


def extract_signals(seed: Seed, pages: list[dict[str, Any]]) -> dict[str, Any]:
    ok_pages = [page for page in pages if page.get("ok_status")]
    combined = clean_text(" ".join(page["markdown"] for page in ok_pages))
    signals: dict[str, Any] = {}
    signal_sources: dict[str, list[dict[str, str]]] = {}
    for key, terms in SIGNAL_TERMS.items():
        details: list[dict[str, str]] = []
        for page in ok_pages:
            details.extend(find_snippet_details(page, terms, limit=2))
            if len(details) >= 4:
                break
        signal_sources[key] = details[:4]
        signals[key] = [detail["snippet"] for detail in signal_sources[key]]

    operation_models = infer_operation(combined)
    if operation_models:
        signals["operation_models"] = operation_models

    availability = infer_availability(ok_pages)
    product_images = extract_product_image_candidates(ok_pages)
    missing = [message for key, message in MISSING_RULES.items() if not signals.get(key)]
    evidence_count = sum(1 for key in ["public_sector", "privacy", "hosting"] if signals.get(key))
    confidence = "hoch" if evidence_count >= 3 else "mittel" if evidence_count >= 2 else "offen"
    signal_payload = {
        "signals": signals,
        "availability": availability,
        "missing_info": missing,
        "confidence": confidence,
    }
    monitor_payload = {
        "signal_presence": {key: bool(signals.get(key)) for key in SIGNAL_TERMS},
        "operation_models": sorted(signals.get("operation_models", [])),
        "availability_scope": availability["scope"],
        "availability_regions": availability["regions"],
        "missing_info": missing,
        "confidence": confidence,
    }

    return {
        "seed": dataclasses.asdict(seed),
        "signals": signals,
        "signal_sources": signal_sources,
        "availability": availability,
        "product_image_candidates": product_images,
        "missing_info": missing,
        "confidence": confidence,
        "content_hash": hashlib.sha256(combined.encode("utf-8")).hexdigest(),
        "signal_hash": hashlib.sha256(
            json.dumps(signal_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest(),
        "monitor_hash": hashlib.sha256(
            json.dumps(monitor_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest(),
        "content_length": len(combined),
    }


def signal_hash_from_record(record: dict[str, Any]) -> str:
    payload = {
        "signals": record.get("signals", {}),
        "availability": record.get("availability", {}),
        "missing_info": record.get("missing_info", []),
        "confidence": record.get("confidence"),
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def monitor_hash_from_record(record: dict[str, Any]) -> str:
    signals = record.get("signals", {})
    payload = {
        "signal_presence": {key: bool(signals.get(key)) for key in SIGNAL_TERMS},
        "operation_models": sorted(signals.get("operation_models", [])),
        "availability_scope": record.get("availability", {}).get("scope"),
        "availability_regions": record.get("availability", {}).get("regions", []),
        "missing_info": record.get("missing_info", []),
        "confidence": record.get("confidence"),
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def compare_state(slug: str, current: dict[str, Any], previous_state: dict[str, Any]) -> dict[str, Any]:
    previous = previous_state.get("products", {}).get(slug)
    if not previous:
        return {"status": "new", "changed_fields": ["first_seen"]}

    changed_fields: list[str] = []
    previous_monitor_hash = previous.get("monitor_hash") or monitor_hash_from_record(previous)
    if previous_monitor_hash != current.get("monitor_hash"):
        changed_fields.append("monitor_signals")
    if previous.get("confidence") != current.get("confidence"):
        changed_fields.append("confidence")
    if previous.get("missing_info") != current.get("missing_info"):
        changed_fields.append("missing_info")
    if previous.get("availability", {}).get("scope") != current.get("availability", {}).get("scope"):
        changed_fields.append("availability_scope")
    if previous.get("availability", {}).get("regions", []) != current.get("availability", {}).get("regions", []):
        changed_fields.append("availability_regions")

    old_signals = previous.get("signals", {})
    new_signals = current.get("signals", {})
    for key in sorted(set(old_signals) | set(new_signals)):
        if bool(old_signals.get(key)) != bool(new_signals.get(key)):
            changed_fields.append(f"signal:{key}")

    return {
        "status": "changed" if changed_fields else "unchanged",
        "changed_fields": changed_fields,
    }


def suggested_decision(item: dict[str, Any]) -> str:
    if item["crawl"]["successful_pages"] == 0:
        return "Nachrecherche"
    if item["missing_info"]:
        return "Nachrecherche"
    if item["confidence"] == "hoch":
        return "Kandidat fuer Uebernahme"
    return "Pruefen"


def write_raw(raw_dir: Path, seed: Seed, pages: list[dict[str, Any]]) -> None:
    raw_dir.mkdir(parents=True, exist_ok=True)
    parts = [f"# {seed.company}", ""]
    for page in pages:
        parts.extend(
            [
                f"## {page['url']}",
                f"success: {page['success']}",
                f"status: {page.get('status_code')}",
                f"ok_status: {page.get('ok_status', False)}",
                f"error: {page.get('error') or ''}",
                "",
                (page.get("markdown") or "")[:60000],
                "",
            ]
        )
    (raw_dir / f"{slugify(seed.company)}.md").write_text("\n".join(parts), encoding="utf-8")


def truncate(value: str, length: int = 260) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    if len(value) <= length:
        return value
    return value[: length - 1].rstrip() + "..."


def format_signal_sources(sources: list[dict[str, str]], limit: int = 2) -> list[str]:
    if not sources:
        return ["nicht gefunden"]
    lines = []
    for source in sources[:limit]:
        lines.append(f"{source['url']} - {truncate(source['snippet'])}")
    return lines


def yes_no(value: bool) -> str:
    return "ja" if value else "nein"


def format_snippets(snippets: list[str]) -> str:
    if not snippets:
        return "nicht gefunden"
    return "; ".join(truncate(snippet, 220) for snippet in snippets[:2])


def build_report(run_meta: dict[str, Any], candidates: list[dict[str, Any]], changes: dict[str, Any]) -> str:
    warnings = [
        item
        for item in candidates
        if item["crawl"]["successful_pages"] == 0 or item["missing_info"] or item["change"]["status"] == "changed"
    ]
    strong_candidates = [
        item
        for item in candidates
        if item["confidence"] == "hoch" and not item["missing_info"] and item["crawl"]["successful_pages"] > 0
    ]

    lines = [
        "# Supertools Crawler Review",
        "",
        f"Run: `{run_meta['run_id']}`",
        f"Zeit: {run_meta['started_at']}",
        f"Seed-Datei: `{run_meta['excel']}`",
        f"Anbieter im Lauf: {len(candidates)}",
        "",
        "## Zusammenfassung",
        "",
        f"- Neu: {changes['summary']['new']}",
        f"- Geaendert: {changes['summary']['changed']}",
        f"- Unveraendert: {changes['summary']['unchanged']}",
        f"- Fehlerhaft gecrawlt: {changes['summary']['failed']}",
        "",
        "## Entscheidungstabelle",
        "",
        "| Anbieter | Status | Vorschlag | Confidence | Seiten | Fehlende Infos / Content / Bilder |",
        "| --- | --- | --- | --- | ---: | ---: |",
    ]

    for item in candidates:
        seed = item["seed"]
        lines.append(
            "| "
            + " | ".join(
                [
                    seed["company"].replace("|", "\\|"),
                    f"`{item['change']['status']}`",
                    suggested_decision(item),
                    f"`{item['confidence']}`",
                    f"{item['crawl']['successful_pages']}/{item['crawl']['attempted_pages']}",
                    f"{len(item['missing_info'])} / Content {len(item.get('content_pieces', []))} / Bilder {len(item.get('product_image_candidates', []))}",
                ]
            )
            + " |"
        )

    lines.extend(["", "## Sofort Pruefen", ""])
    if warnings:
        for item in warnings:
            seed = item["seed"]
            reasons = []
            if item["crawl"]["successful_pages"] == 0:
                reasons.append("keine Seite erfolgreich gecrawlt")
            if item["missing_info"]:
                reasons.append(f"{len(item['missing_info'])} fehlende Pflichtinfos")
            if item["change"]["status"] == "changed":
                reasons.append("Monitoring-Signal geaendert")
            lines.append(f"- {seed['company']}: {', '.join(reasons)}")
    else:
        lines.append("- Keine harten Warnungen in diesem Lauf.")

    lines.extend(["", "## Gute Kandidaten fuer die naechste Qualifizierung", ""])
    if strong_candidates:
        for item in strong_candidates:
            seed = item["seed"]
            lines.append(f"- {seed['company']} ({seed['branch']})")
    else:
        lines.append("- Keine Kandidaten ohne MVP-Pflichtluecke.")

    lines.extend(["", "## Review-Liste", ""])

    for item in candidates:
        seed = item["seed"]
        signals = item["signals"]
        signal_sources = item.get("signal_sources", {})
        crawl = item["crawl"]
        content_pieces = item.get("content_pieces", [])
        product_images = item.get("product_image_candidates", [])
        status = item["change"]["status"]
        operation_summary = ", ".join(signals.get("operation_models", [])) or format_snippets(signals.get("operation", []))
        availability = item.get("availability", {})
        availability_regions = ", ".join(availability.get("regions", []))
        availability_summary = availability.get("label") or "unklar"
        if availability_regions:
            availability_summary = f"{availability_summary}: {availability_regions}"
        availability_summary = f"{availability_summary} ({availability.get('confidence', 'offen')})"
        present = {key: bool(signals.get(key)) for key in SIGNAL_LABELS}
        source_urls = [entry["url"] for entry in crawl["urls"]]
        failed_urls = [entry for entry in crawl["urls"] if not entry.get("ok_status")]

        lines.extend(
            [
                f"### {seed['company']}",
                "",
                f"Vorschlag: **{suggested_decision(item)}**",
                "",
                "| Feld | Wert |",
                "| --- | --- |",
                f"| Status | `{status}` |",
                f"| Website | {seed['website']} |",
                f"| Cluster / interner Score | {seed['cluster']} / {seed['relevance_score']} |",
                f"| Branche | {seed['branch']} |",
                f"| Hauptsitz / Mitarbeiter | {seed['city'] or 'offen'} / {seed['employees'] or 'offen'} |",
                f"| Review-Confidence | `{item['confidence']}` |",
                f"| Crawling | {crawl['successful_pages']}/{crawl['attempted_pages']} Seiten erfolgreich |",
                f"| Betriebsmodell | {operation_summary} |",
                f"| Verfuegbarkeit | {availability_summary} |",
                f"| Content Pieces | {len(content_pieces)} gefunden |",
                f"| Produktbild-Kandidaten | {len(product_images)} gefunden; immer redaktionell pruefen |",
                "",
                "**Signalampel**",
                "",
                "| Signal | Gefunden |",
                "| --- | --- |",
            ]
        )

        for key, label in SIGNAL_LABELS.items():
            lines.append(f"| {label} | {yes_no(present[key])} |")
        lines.append(f"| Verfuegbarkeit klaerbar | {yes_no(bool(availability.get('evidence')))} |")

        lines.extend(["", "**Gecrawlte Quellen**", ""])
        lines.extend(f"- {url}" for url in source_urls)

        if failed_urls:
            lines.extend(["", "**Crawl-Probleme**", ""])
            for entry in failed_urls:
                lines.append(
                    f"- {entry['url']} - Status {entry.get('status_code')}, {entry.get('error') or 'kein verwertbarer Inhalt'}"
                )

        lines.extend(["", "**Quellen-Signale**", ""])
        for key, label in SIGNAL_LABELS.items():
            lines.append(f"- {label}:")
            for detail in format_signal_sources(signal_sources.get(key, [])):
                lines.append(f"  - {detail}")
        lines.append("- Verfuegbarkeit:")
        if availability.get("evidence"):
            for detail in availability["evidence"][:3]:
                region = f" [{detail.get('region')}]" if detail.get("region") else ""
                lines.append(f"  - {detail['url']}{region} - {truncate(detail['snippet'])}")
        else:
            lines.append("  - nicht gefunden")

        lines.extend(["", "**Fehlende Informationen**", ""])
        if item["missing_info"]:
            lines.extend(f"- {entry}" for entry in item["missing_info"])
        else:
            lines.append("- Keine MVP-Pflichtluecke erkannt.")

        lines.extend(["", "**Produktbild-Kandidaten (keine Freigabe)**", ""])
        if product_images:
            for image in product_images[:8]:
                lines.append(f"- Score {image['score']}: {image['url']}")
                lines.append(f"  - Fundstelle: {image['source_url']}")
                if image.get("alt"):
                    lines.append(f"  - Alt/Text: {truncate(image['alt'], 180)}")
                if image.get("reasons"):
                    lines.append(f"  - Signale: {', '.join(image['reasons'])}")
        else:
            lines.append("- Keine Kandidaten gefunden; Produktbild beim Anbieter anfragen oder gezielt nachrecherchieren.")

        lines.extend(["", "**Gefundene Content Pieces**", ""])
        if content_pieces:
            for piece in content_pieces[:8]:
                meta = []
                if piece.get("platform"):
                    meta.append(str(piece["platform"]))
                if piece.get("video_id"):
                    meta.append(f"video_id={piece['video_id']}")
                lines.append(f"- `{piece['kind']}` {piece['title']} - {piece['url']}")
                lines.append(f"  - Quelle: {piece['source_url']}")
                if meta:
                    lines.append(f"  - Meta: {', '.join(meta)}")
                if piece.get("thumbnail_url"):
                    lines.append(f"  - Thumbnail: {piece['thumbnail_url']}")
        else:
            lines.append("- Keine passenden Videos, Webinare, Cases, Whitepaper, Blogartikel oder PDFs gefunden.")

        lines.extend(
            [
                "",
                "**Redaktionelle Entscheidung**",
                "",
                "- [ ] Uebernehmen",
                "- [ ] Nachrecherche",
                "- [ ] Anbieter kontaktieren",
                "- [ ] Ablehnen / ignorieren",
                "",
            ]
        )

    lines.extend(
        [
            "",
            "## Hinweise zur Nutzung",
            "",
            "- Dieser Report ist eine Vorqualifizierung, keine Veroeffentlichungsfreigabe.",
            "- Snippets muessen vor Uebernahme redaktionell gegengeprueft werden.",
            "- Fehlende Informationen sollen spaeter sichtbar gemacht oder beim Anbieter nachgefragt werden.",
        ]
    )

    return "\n".join(lines)


async def process_seed(crawler: AsyncWebCrawler, seed: Seed, max_pages: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    homepage = await crawl_url(crawler, seed.website)
    urls = [seed.website]
    if homepage.get("ok_status"):
        urls.extend(discover_priority_links(seed.website, homepage["html"], max_pages))

    seen: set[str] = set()
    pages = []
    for url in urls[:max_pages]:
        if url in seen:
            continue
        seen.add(url)
        page = homepage if url == seed.website else await crawl_url(crawler, url)
        pages.append(page)

    extracted = extract_signals(seed, pages)
    extracted["content_pieces"] = extract_content_pieces(pages)
    extracted["crawl"] = {
        "attempted_pages": len(pages),
        "successful_pages": sum(1 for page in pages if page.get("ok_status")),
        "urls": [
            {
                "url": page["url"],
                "success": page["success"],
                "status_code": page.get("status_code"),
                "ok_status": page.get("ok_status", False),
                "error": page.get("error") or "",
            }
            for page in pages
        ],
    }
    return extracted, pages


async def run(args: argparse.Namespace) -> None:
    started = dt.datetime.now(dt.timezone.utc).astimezone()
    run_id = args.run_id or started.strftime("%Y%m%d-%H%M%S")
    output_dir = Path(args.out) / run_id
    raw_dir = output_dir / "raw"
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_seeds = [] if args.seed_source == "curation" else read_seeds(Path(args.excel), args.cluster, args.score_min)
    curation_seeds = (
        []
        if args.seed_source == "excel"
        else read_curation_seeds(Path(args.curation_seeds), args.cluster, args.score_min)
    )
    seeds = merge_seeds(excel_seeds, curation_seeds)
    seeds = seeds[args.offset : args.offset + args.limit]
    state_path = Path(args.state)
    state = load_state(state_path)
    new_state = json.loads(json.dumps(state))
    new_state.setdefault("products", {})

    browser_config = BrowserConfig(
        headless=True,
        viewport_width=1440,
        viewport_height=1000,
        text_mode=False,
        verbose=False,
    )

    candidates: list[dict[str, Any]] = []
    changes = {
        "summary": {"new": 0, "changed": 0, "unchanged": 0, "failed": 0},
        "items": [],
    }

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for index, seed in enumerate(seeds, start=1):
            print(f"[{index}/{len(seeds)}] Crawling {seed.company} - {seed.website}")
            extracted, pages = await process_seed(crawler, seed, args.pages_per_company)
            slug = slugify(seed.company)
            change = compare_state(slug, extracted, state)
            extracted["slug"] = slug
            extracted["change"] = change
            extracted["review_status"] = "needs_review"
            extracted["last_checked_at"] = started.date().isoformat()
            candidates.append(extracted)
            write_raw(raw_dir, seed, pages)

            changes["summary"][change["status"]] += 1
            if not any(page.get("ok_status") for page in pages):
                changes["summary"]["failed"] += 1
            changes["items"].append(
                {
                    "slug": slug,
                    "company": seed.company,
                    "website": seed.website,
                    "status": change["status"],
                    "changed_fields": change["changed_fields"],
                    "successful_pages": extracted["crawl"]["successful_pages"],
                }
            )

            new_state["products"][slug] = {
                "company": seed.company,
                "website": seed.website,
                "last_checked_at": extracted["last_checked_at"],
                "content_hash": extracted["content_hash"],
                "signal_hash": extracted["signal_hash"],
                "monitor_hash": extracted["monitor_hash"],
                "confidence": extracted["confidence"],
                "missing_info": extracted["missing_info"],
                "signals": extracted["signals"],
                "availability": extracted["availability"],
                "crawl": extracted["crawl"],
            }

    run_meta = {
        "run_id": run_id,
        "started_at": started.isoformat(timespec="seconds"),
        "excel": args.excel,
        "curation_seeds": args.curation_seeds,
        "seed_source": args.seed_source,
        "excel_seed_count": len(excel_seeds),
        "curation_seed_count": len(curation_seeds),
        "limit": args.limit,
        "offset": args.offset,
        "cluster": args.cluster,
        "score_min": args.score_min,
        "pages_per_company": args.pages_per_company,
        "saved_state": not args.no_save_state,
    }
    save_json(output_dir / "run-meta.json", run_meta)
    save_json(output_dir / "product-candidates.json", candidates)
    save_json(output_dir / "changes.json", changes)
    (output_dir / "review-report.md").write_text(build_report(run_meta, candidates, changes), encoding="utf-8")

    if not args.no_save_state:
        save_json(state_path, new_state)

    print("")
    print(f"Wrote {output_dir / 'review-report.md'}")
    print(f"Wrote {output_dir / 'product-candidates.json'}")
    print(f"Wrote {output_dir / 'changes.json'}")
    print(f"State {'not updated' if args.no_save_state else 'updated'}: {state_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--curation-seeds", default=DEFAULT_CURATION_SEEDS)
    parser.add_argument("--seed-source", choices=["all", "excel", "curation"], default="all")
    parser.add_argument("--state", default=DEFAULT_STATE)
    parser.add_argument("--out", default=DEFAULT_RUNS)
    parser.add_argument("--run-id")
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--cluster")
    parser.add_argument("--score-min", type=int)
    parser.add_argument("--pages-per-company", type=int, default=4)
    parser.add_argument("--no-save-state", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    asyncio.run(run(parse_args()))
